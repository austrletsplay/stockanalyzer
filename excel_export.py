"""
excel_export.py – Erstellt eine formatierte Excel-Datei mit Fundamentalanalyse.
"""

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# ── Farben ────────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SUBHEADER_FILL = PatternFill("solid", fgColor="2E75B6")
ALT_FILL = PatternFill("solid", fgColor="F2F9FF")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
GREY_FILL = PatternFill("solid", fgColor="F2F2F2")
BUY_FILL = PatternFill("solid", fgColor="1B5E20")
HOLD_FILL = PatternFill("solid", fgColor="BF360C")
SELL_FILL = PatternFill("solid", fgColor="B71C1C")

WHITE_BOLD = Font(bold=True, color="FFFFFF")
DARK_BOLD = Font(bold=True, color="1F2937")
NORMAL = Font(color="1F2937")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
WRAP = Alignment(wrap_text=True, vertical="top")

THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)


def generate_excel(metrics: dict, scores: dict, ticker_symbol: str) -> bytes:
    """
    Erzeugt eine Excel-Datei und gibt sie als Bytes zurück (für st.download_button).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"{ticker_symbol} Analyse"

    _set_column_widths(ws)
    row = 1

    row = _write_header_block(ws, metrics, scores, ticker_symbol, row)
    row = _write_score_overview(ws, scores, row)
    row = _write_metric_details(ws, metrics, scores, row)
    row = _write_historical_data(ws, metrics, row)
    row = _write_report_text(ws, metrics, scores, row)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ──────────────────────────────────────────────────────────────────────────────
# Abschnitt 1: Header
# ──────────────────────────────────────────────────────────────────────────────

def _write_header_block(ws, metrics, scores, ticker, start_row):
    company = metrics['company']
    row = start_row

    # Firmenname
    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row, 1, f"{ticker} — {company['name']}")
    c.fill = HEADER_FILL
    c.font = Font(bold=True, color="FFFFFF", size=14)
    c.alignment = CENTER
    ws.row_dimensions[row].height = 30
    row += 1

    # Sektor / Land / Währung
    ws.merge_cells(f"A{row}:B{row}")
    _write_cell(ws, row, 1, f"Sektor: {company.get('sector', 'N/A')}", fill=GREY_FILL, bold=False)
    _write_cell(ws, row, 3, f"Land: {company.get('country', 'N/A')}", fill=GREY_FILL, bold=False)
    _write_cell(ws, row, 4, f"Währung: {company.get('currency', 'N/A')}", fill=GREY_FILL, bold=False)
    _write_cell(ws, row, 5, f"Market Cap: {company.get('market_cap_fmt', 'N/A')}", fill=GREY_FILL, bold=False)
    row += 1

    _write_cell(ws, row, 1, f"Branche: {company.get('industry', 'N/A')}", fill=GREY_FILL, bold=False)
    _write_cell(ws, row, 4, f"Analysedatum: {date.today().strftime('%d.%m.%Y')}", fill=GREY_FILL, bold=False)
    row += 1

    # Leerzeile
    row += 1

    # Empfehlungs-Block
    rec = scores['recommendation']
    rec_color = {"KAUFEN": "1B5E20", "HALTEN": "E65100", "NICHT KAUFEN": "B71C1C"}.get(rec, "333333")
    rec_fill = PatternFill("solid", fgColor=rec_color)

    ws.merge_cells(f"A{row}:C{row}")
    c = ws.cell(row, 1, f"EMPFEHLUNG: {rec}")
    c.fill = rec_fill
    c.font = Font(bold=True, color="FFFFFF", size=13)
    c.alignment = CENTER
    ws.merge_cells(f"D{row}:F{row}")
    c2 = ws.cell(row, 4, f"Gesamtscore: {scores['total_score']} / 100 Punkte")
    c2.fill = rec_fill
    c2.font = Font(bold=True, color="FFFFFF", size=13)
    c2.alignment = CENTER
    ws.row_dimensions[row].height = 28
    row += 2

    return row


# ──────────────────────────────────────────────────────────────────────────────
# Abschnitt 2: Score-Übersicht
# ──────────────────────────────────────────────────────────────────────────────

def _write_score_overview(ws, scores, start_row):
    row = start_row

    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row, 1, "SCORE-ÜBERSICHT")
    c.fill = SUBHEADER_FILL
    c.font = WHITE_BOLD
    c.alignment = CENTER
    row += 1

    # Tabellenkopf
    headers = ["Kategorie", "Punkte", "Maximum", "Prozent", "Bewertung", ""]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row, col, h)
        c.fill = GREY_FILL
        c.font = DARK_BOLD
        c.alignment = CENTER
        c.border = THIN_BORDER
    row += 1

    categories = [
        ("Wachstum (Growth)", "growth"),
        ("Profitabilität", "profitability"),
        ("Bilanzqualität", "balance_sheet"),
        ("Bewertung (Valuation)", "valuation"),
    ]

    for i, (label, key) in enumerate(categories):
        cat = scores['category_scores'][key]
        score = cat['score']
        max_pts = cat['max']
        pct = round(score / max_pts * 100) if max_pts > 0 else 0
        bar = "█" * (score * 10 // max_pts) + "░" * (10 - score * 10 // max_pts)
        fill = ALT_FILL if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        score_fill = GREEN_FILL if pct >= 70 else (YELLOW_FILL if pct >= 45 else RED_FILL)

        _write_cell(ws, row, 1, label, fill=fill)
        _write_cell(ws, row, 2, score, fill=score_fill, align=CENTER)
        _write_cell(ws, row, 3, max_pts, fill=fill, align=CENTER)
        _write_cell(ws, row, 4, f"{pct}%", fill=fill, align=CENTER)
        _write_cell(ws, row, 5, bar, fill=fill, align=CENTER)
        row += 1

    # Gesamt-Zeile
    total = scores['total_score']
    total_pct = total
    total_fill = GREEN_FILL if total >= 75 else (YELLOW_FILL if total >= 50 else RED_FILL)
    _write_cell(ws, row, 1, "GESAMT", bold=True, fill=GREY_FILL)
    _write_cell(ws, row, 2, total, bold=True, fill=total_fill, align=CENTER)
    _write_cell(ws, row, 3, 100, bold=True, fill=GREY_FILL, align=CENTER)
    _write_cell(ws, row, 4, f"{total_pct}%", bold=True, fill=total_fill, align=CENTER)
    row += 2

    return row


# ──────────────────────────────────────────────────────────────────────────────
# Abschnitt 3: Detailtabellen je Kategorie
# ──────────────────────────────────────────────────────────────────────────────

def _write_metric_details(ws, metrics, scores, start_row):
    row = start_row
    g = metrics['growth']
    p = metrics['profitability']
    b = metrics['balance_sheet']
    v = metrics['valuation']

    def fmt(val, pct=False, mult=False, x=False):
        if val is None:
            return "N/A"
        if pct:
            return f"{val*100:.1f}%"
        if x:
            return f"{val:.2f}x"
        if mult:
            return f"{val:.2f}x"
        return f"{val:.2f}"

    # ── WACHSTUM ──
    row = _write_category_header(ws, row, "WACHSTUM (GROWTH)", "growth", scores)
    row = _write_detail_table_header(ws, row)
    growth_rows = [
        ("Umsatz-CAGR (3 Jahre)", fmt(g.get('revenue_cagr_3yr'), pct=True),
         "≥15%", scores['category_scores']['growth']['breakdown'].get('revenue_cagr', {})),
        ("EPS-CAGR (3 Jahre)", fmt(g.get('eps_cagr_3yr'), pct=True),
         "≥15%", scores['category_scores']['growth']['breakdown'].get('eps_cagr', {})),
        ("EPS-Konsistenz (YoY pos.)", _fmt_consistency(g.get('eps_consistency')),
         "100%", scores['category_scores']['growth']['breakdown'].get('eps_consistency', {})),
        ("Umsatzwachstum YoY", fmt(g.get('revenue_yoy'), pct=True), "≥10%", {}),
        ("Gewinnwachstum YoY", fmt(g.get('earnings_yoy'), pct=True), "≥10%", {}),
    ]
    for r in growth_rows:
        row = _write_detail_row(ws, row, *r)
    row += 1

    # ── PROFITABILITÄT ──
    row = _write_category_header(ws, row, "PROFITABILITÄT", "profitability", scores)
    row = _write_detail_table_header(ws, row)
    prof_rows = [
        ("ROE (Eigenkapitalrendite)", fmt(p.get('roe'), pct=True),
         "≥15%", scores['category_scores']['profitability']['breakdown'].get('roe', {})),
        ("Nettomarge", fmt(p.get('net_margin'), pct=True),
         "≥10%", scores['category_scores']['profitability']['breakdown'].get('net_margin', {})),
        ("FCF-Marge", fmt(p.get('fcf_margin'), pct=True),
         "≥8%", scores['category_scores']['profitability']['breakdown'].get('fcf_margin', {})),
        ("Bruttomarge", fmt(p.get('gross_margin'), pct=True), "≥30%", {}),
        ("EBIT-Marge", fmt(p.get('operating_margin'), pct=True), "≥15%", {}),
        ("EBITDA-Marge", fmt(p.get('ebitda_margin'), pct=True), "≥20%", {}),
        ("ROA (Gesamtkapitalrendite)", fmt(p.get('roa'), pct=True), "≥5%", {}),
    ]
    for r in prof_rows:
        row = _write_detail_row(ws, row, *r)
    row += 1

    # ── BILANZQUALITÄT ──
    row = _write_category_header(ws, row, "BILANZQUALITÄT", "balance_sheet", scores)
    row = _write_detail_table_header(ws, row)
    bal_rows = [
        ("Debt/Equity Ratio", fmt(b.get('debt_to_equity'), x=True),
         "<0.5x", scores['category_scores']['balance_sheet']['breakdown'].get('debt_equity', {})),
        ("Current Ratio", fmt(b.get('current_ratio')),
         "≥1.5", scores['category_scores']['balance_sheet']['breakdown'].get('current_ratio', {})),
        ("Quick Ratio", fmt(b.get('quick_ratio')), "≥1.0", {}),
        ("FCF-Konsistenz", _fmt_fcf_consistency(b.get('fcf_history', [])),
         "4/4 pos.", scores['category_scores']['balance_sheet']['breakdown'].get('fcf_consistency', {})),
        ("Gesamtschulden", _fmt_number(b.get('total_debt')), "—", {}),
        ("Kassenbestand", _fmt_number(b.get('total_cash')), "—", {}),
        ("Nettoverschuldung", _fmt_number(b.get('net_debt')), "Negativ = Netto-Cash", {}),
    ]
    for r in bal_rows:
        row = _write_detail_row(ws, row, *r)
    row += 1

    # ── BEWERTUNG ──
    row = _write_category_header(ws, row, "BEWERTUNG (VALUATION)", "valuation", scores)
    row = _write_detail_table_header(ws, row)
    val_rows = [
        ("PEG-Ratio", fmt(v.get('peg')),
         "<1.5", scores['category_scores']['valuation']['breakdown'].get('peg', {})),
        ("P/FCF (Kurs/FCF)", fmt(v.get('p_fcf'), x=True),
         "<25x", scores['category_scores']['valuation']['breakdown'].get('p_fcf', {})),
        ("Forward KGV (P/E)", fmt(v.get('forward_pe'), x=True),
         "<25x", scores['category_scores']['valuation']['breakdown'].get('forward_pe', {})),
        ("Trailing KGV (P/E)", fmt(v.get('pe'), x=True), "<30x", {}),
        ("KBV (P/Book)", fmt(v.get('pb'), x=True), "<3x", {}),
        ("KUV (P/Sales)", fmt(v.get('ps'), x=True), "<5x", {}),
        ("EV/EBITDA", fmt(v.get('ev_ebitda'), x=True), "<15x", {}),
    ]
    for r in val_rows:
        row = _write_detail_row(ws, row, *r)
    row += 1

    return row


def _write_category_header(ws, row, title, key, scores):
    cat = scores['category_scores'][key]
    score = cat['score']
    max_pts = cat['max']
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row, 1, title)
    c.fill = SUBHEADER_FILL
    c.font = WHITE_BOLD
    c.alignment = LEFT
    ws.merge_cells(f"E{row}:F{row}")
    c2 = ws.cell(row, 5, f"{score} / {max_pts} Punkte")
    c2.fill = SUBHEADER_FILL
    c2.font = WHITE_BOLD
    c2.alignment = CENTER
    ws.row_dimensions[row].height = 22
    return row + 1


def _write_detail_table_header(ws, row):
    cols = ["Kennzahl", "Wert", "Benchmark", "Punkte", "Bewertung", ""]
    for col, h in enumerate(cols, 1):
        c = ws.cell(row, col, h)
        c.fill = GREY_FILL
        c.font = DARK_BOLD
        c.alignment = CENTER
        c.border = THIN_BORDER
    return row + 1


def _write_detail_row(ws, row, label, value, benchmark, breakdown: dict):
    pts = breakdown.get('pts')
    max_pts = breakdown.get('max')
    detail_label = breakdown.get('label', '')

    pts_str = f"{pts}/{max_pts}" if pts is not None and max_pts is not None else "—"

    # Zellfärbung je nach Score
    if pts is not None and max_pts and max_pts > 0:
        ratio = pts / max_pts
        score_fill = GREEN_FILL if ratio >= 0.7 else (YELLOW_FILL if ratio >= 0.4 else RED_FILL)
    else:
        score_fill = PatternFill("solid", fgColor="FFFFFF")

    fill = ALT_FILL if row % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

    _write_cell(ws, row, 1, label, fill=fill)
    _write_cell(ws, row, 2, value, fill=fill, align=CENTER)
    _write_cell(ws, row, 3, benchmark, fill=fill, align=CENTER)
    _write_cell(ws, row, 4, pts_str, fill=score_fill, align=CENTER, bold=pts is not None)
    _write_cell(ws, row, 5, detail_label, fill=fill)
    return row + 1


# ──────────────────────────────────────────────────────────────────────────────
# Abschnitt 4: Historische Daten
# ──────────────────────────────────────────────────────────────────────────────

def _write_historical_data(ws, metrics, start_row):
    row = start_row
    g = metrics['growth']
    b = metrics['balance_sheet']
    company = metrics['company']
    currency = company.get('currency', 'USD')

    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row, 1, "HISTORISCHE DATEN")
    c.fill = SUBHEADER_FILL
    c.font = WHITE_BOLD
    c.alignment = CENTER
    row += 1

    years = g.get('revenue_years') or b.get('fcf_years') or []
    rev_hist = g.get('revenue_history', [])
    eps_hist = g.get('eps_history', [])
    fcf_hist = b.get('fcf_history', [])

    if not years:
        _write_cell(ws, row, 1, "Keine historischen Daten verfügbar")
        return row + 2

    # Tabellenkopf
    headers = ["Jahr", f"Umsatz ({currency})", "EPS", f"FCF ({currency})", "", ""]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row, col, h)
        c.fill = GREY_FILL
        c.font = DARK_BOLD
        c.alignment = CENTER
        c.border = THIN_BORDER
    row += 1

    max_len = max(len(years), len(rev_hist), len(eps_hist), len(fcf_hist))
    for i in range(max_len):
        year = years[i] if i < len(years) else "—"
        rev = _fmt_number(rev_hist[i]) if i < len(rev_hist) else "—"
        eps = f"{eps_hist[i]:.2f}" if i < len(eps_hist) else "—"
        fcf = _fmt_number(fcf_hist[i]) if i < len(fcf_hist) else "—"

        fill = ALT_FILL if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        _write_cell(ws, row, 1, year, fill=fill, align=CENTER)
        _write_cell(ws, row, 2, rev, fill=fill, align=CENTER)
        _write_cell(ws, row, 3, eps, fill=fill, align=CENTER)
        _write_cell(ws, row, 4, fcf, fill=fill, align=CENTER)
        row += 1

    return row + 1


# ──────────────────────────────────────────────────────────────────────────────
# Abschnitt 5: Report-Text
# ──────────────────────────────────────────────────────────────────────────────

def _write_report_text(ws, metrics, scores, start_row):
    row = start_row

    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row, 1, "ANALYSEBERICHT")
    c.fill = SUBHEADER_FILL
    c.font = WHITE_BOLD
    c.alignment = CENTER
    row += 1

    # Stärken
    if scores['strengths']:
        _write_cell(ws, row, 1, "Stärken:", bold=True, fill=GREEN_FILL)
        row += 1
        for s in scores['strengths']:
            ws.merge_cells(f"A{row}:F{row}")
            c = ws.cell(row, 1, f"  ✓  {s}")
            c.fill = PatternFill("solid", fgColor="E8F5E9")
            c.alignment = WRAP
            ws.row_dimensions[row].height = 20
            row += 1
        row += 1

    # Risiken
    if scores['concerns']:
        _write_cell(ws, row, 1, "Risiken / Schwächen:", bold=True, fill=RED_FILL)
        row += 1
        for concern in scores['concerns']:
            ws.merge_cells(f"A{row}:F{row}")
            c = ws.cell(row, 1, f"  ⚠  {concern}")
            c.fill = PatternFill("solid", fgColor="FFEBEE")
            c.alignment = WRAP
            ws.row_dimensions[row].height = 20
            row += 1
        row += 1

    # Datenwarnungen
    if scores['data_warnings']:
        _write_cell(ws, row, 1, "Datenverfügbarkeit:", bold=True, fill=YELLOW_FILL)
        row += 1
        for w in scores['data_warnings']:
            ws.merge_cells(f"A{row}:F{row}")
            c = ws.cell(row, 1, f"  ℹ  {w}")
            c.fill = PatternFill("solid", fgColor="FFFDE7")
            c.alignment = WRAP
            ws.row_dimensions[row].height = 20
            row += 1
        row += 1

    # Unternehmensbeschreibung
    desc = metrics['company'].get('description', '')
    if desc:
        _write_cell(ws, row, 1, "Unternehmensübersicht:", bold=True, fill=GREY_FILL)
        row += 1
        ws.merge_cells(f"A{row}:F{row + 3}")
        c = ws.cell(row, 1, desc[:1000])  # Max 1000 Zeichen
        c.alignment = WRAP
        c.fill = PatternFill("solid", fgColor="FAFAFA")
        ws.row_dimensions[row].height = 80
        row += 5

    # Disclaimer
    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row, 1,
        "⚠ Hinweis: Diese Analyse dient nur zu Informationszwecken und stellt keine Anlageberatung dar. "
        "Alle Investitionsentscheidungen sollten auf eigener Recherche basieren.")
    c.fill = YELLOW_FILL
    c.font = Font(italic=True, size=9, color="333333")
    c.alignment = WRAP
    ws.row_dimensions[row].height = 30
    row += 1

    return row


# ──────────────────────────────────────────────────────────────────────────────
# Hilfsfunktionen
# ──────────────────────────────────────────────────────────────────────────────

def _set_column_widths(ws):
    widths = {'A': 30, 'B': 16, 'C': 16, 'D': 14, 'E': 35, 'F': 14}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _write_cell(ws, row, col, value, fill=None, bold=False, align=None, border=True):
    c = ws.cell(row, col, value)
    if fill:
        c.fill = fill
    if bold:
        c.font = Font(bold=True, color="1F2937")
    else:
        c.font = Font(color="1F2937")
    c.alignment = align or LEFT
    if border:
        c.border = THIN_BORDER
    return c


def _fmt_number(val) -> str:
    if val is None:
        return "N/A"
    try:
        v = float(val)
        if abs(v) >= 1e12:
            return f"{v/1e12:.2f}T"
        elif abs(v) >= 1e9:
            return f"{v/1e9:.2f}B"
        elif abs(v) >= 1e6:
            return f"{v/1e6:.2f}M"
        else:
            return f"{v:,.0f}"
    except (TypeError, ValueError):
        return "N/A"


def _fmt_consistency(val) -> str:
    if val is None:
        return "N/A"
    return f"{val*100:.0f}%"


def _fmt_fcf_consistency(fcf_hist: list) -> str:
    if not fcf_hist:
        return "N/A"
    positive = sum(1 for v in fcf_hist if v > 0)
    return f"{positive}/{len(fcf_hist)} Jahre"
